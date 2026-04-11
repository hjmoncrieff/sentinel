#!/usr/bin/env python3
"""
SENTINEL — Country-Year Join Table
Joins V-Dem, World Bank, and M3 data into a single country × year flat table.

Outputs:
  data/cleaned/country_year.json
  data/cleaned/country_year.csv

Usage:
  python3 scripts/build_country_year.py
  python3 scripts/build_country_year.py --output-dir data/cleaned
"""

import argparse
import csv
import json
from collections import defaultdict
from datetime import UTC, datetime
from pathlib import Path

from openpyxl import load_workbook

ROOT    = Path(__file__).parent.parent
VDEM_IN = ROOT / "data" / "cleaned" / "vdem.json"
WB_IN   = ROOT / "data" / "cleaned" / "worldbank.json"
M3_IN   = ROOT / "data" / "raw" / "M3-Dataset-V1.xlsx"
EVENTS_IN = ROOT / "data" / "canonical" / "events_actor_coded.json"
OUT_DIR = ROOT / "data" / "cleaned"

# ISO3 lookup
ISO3 = {
    "Brazil":"BRA","Colombia":"COL","Mexico":"MEX","Venezuela":"VEN","Argentina":"ARG",
    "Peru":"PER","Chile":"CHL","Ecuador":"ECU","Bolivia":"BOL","Cuba":"CUB",
    "Honduras":"HND","Guatemala":"GTM","El Salvador":"SLV","Nicaragua":"NIC",
    "Paraguay":"PRY","Uruguay":"URY","Haiti":"HTI","Dominican Republic":"DOM",
    "Panama":"PAN","Costa Rica":"CRI","Jamaica":"JAM","Trinidad and Tobago":"TTO",
    "Guyana":"GUY","Suriname":"SUR","Belize":"BLZ","Regional":"REG",
}

LATAM_COUNTRIES = {
    country for country in ISO3
    if country != "Regional"
}

# WB series fields to include (field_name → output column name)
WB_SERIES = {
    "military_expenditure_pct_gdp_series":      "mil_exp_pct_gdp",
    "military_expenditure_current_usd_series":  "mil_exp_usd",
    "military_personnel_total_series":          "mil_personnel",
    "wgi_rule_of_law_series":                   "wgi_rule_of_law",
    "wgi_govt_effectiveness_series":            "wgi_govt_effectiveness",
    "wgi_control_of_corruption_series":         "wgi_control_corruption",
    "wgi_political_stability_series":           "wgi_political_stability",
    "gdp_constant_2015_usd_series":             "gdp_const_2015_usd",
    "gdp_per_capita_constant_2015_usd_series":  "gdp_per_capita_const_usd",
    "population_total_series":                  "population",
    "inflation_consumer_prices_pct_series":     "inflation_consumer_prices_pct",
    "real_interest_rate_series":                "real_interest_rate",
    "official_exchange_rate_series":            "official_exchange_rate",
    "fdi_net_inflows_pct_gdp_series":           "fdi_net_inflows_pct_gdp",
    "debt_service_pct_exports_series":          "debt_service_pct_exports",
    "current_account_pct_gdp_series":           "current_account_pct_gdp",
    "reserves_months_imports_series":           "reserves_months_imports",
    "resource_rents_pct_gdp_series":            "resource_rents_pct_gdp",
    "trade_openness_pct_gdp_series":            "trade_openness_pct_gdp",
    "oda_received_pct_gni_series":              "oda_received_pct_gni",
}

# V-Dem series fields to include
VDEM_SERIES = [
    "polyarchy", "liberal_democracy", "participatory_democracy",
    "deliberative_democracy", "egalitarian_democracy", "regime_type",
    "physinteg", "mil_constrain", "mil_exec", "exec_confidence",
    "judicial_constraints", "legislative_constraints", "rule_of_law_vdem",
    "public_sector_corruption", "executive_corruption", "corruption_index",
    "clientelism", "civil_society_participation", "party_institutionalization",
    "state_authority", "coup_total_events", "coup_event", "coup_attempts",
    "executive_direct_election", "election_repression", "voter_turnout",
    "democracy_breakdown", "democracy_transition", "polity2", "cs_repress",
    "political_violence",
]

M3_FIELDS = [
    "com_mil_serv",
    "com_mil_serv_gen",
    "com_mil_serv_dur_max",
    "com_mil_serv_dur_min",
    "alt_civ_serv",
    "mil_origin",
    "mil_leader",
    "mil_mod",
    "mil_veto",
    "mil_repress",
    "mil_repress_count",
    "mil_impun",
    "milpol_crime",
    "milpol_law",
    "milpol_peace",
    "mil_police",
    "mil_eco_dummy",
    "mil_eco_own",
    "mil_eco_share",
    "mil_eco_dom",
    "milex_gdp",
    "milex_healthexp",
    "pers_to_pop",
    "pers_to_phy",
    "reserve_pop",
    "hwi",
]

DOMESTIC_MILITARY_ROLE_SUBCATEGORIES = {
    "domestic_security_militarization",
    "security_governance_reconfiguration",
    "exceptional_rule_and_coercive_governance",
    "protest_repression_and_security_response",
    "coercive_internal_crackdown",
    "coercive_detention_and_internal_crackdown",
    "security_force_labor_and_institutional_contention",
    "institutional_security_reordering",
    "command_and_coercive_control",
}

MILITARY_POLICING_ROLE_SUBCATEGORIES = {
    "domestic_security_militarization",
    "protest_repression_and_security_response",
    "coercive_internal_crackdown",
    "coercive_detention_and_internal_crackdown",
    "security_force_labor_and_institutional_contention",
}

EXCEPTION_RULE_SUBCATEGORIES = {
    "exceptional_rule_and_coercive_governance",
}


def coerce_missing(value):
    if value in (None, "", "NA", "N/A", -888, "-888"):
        return None
    return value


def parse_m3_numeric(value):
    value = coerce_missing(value)
    if value is None:
        return None
    try:
        numeric = float(value)
    except (TypeError, ValueError):
        return None
    if numeric.is_integer():
        return int(numeric)
    return numeric


def load_m3_rows() -> dict[str, dict[int, dict]]:
    if not M3_IN.exists():
        return {}
    workbook = load_workbook(M3_IN, read_only=True, data_only=True)
    worksheet = workbook[workbook.sheetnames[0]]
    header = [cell for cell in next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True))]
    index = {name: pos for pos, name in enumerate(header)}
    rows_by_country: dict[str, dict[int, dict]] = defaultdict(dict)
    for raw_row in worksheet.iter_rows(min_row=2, values_only=True):
        country = raw_row[index["countryname_standard"]]
        if country not in LATAM_COUNTRIES:
            continue
        year = int(raw_row[index["year"]])
        row = {
            "year": year,
            "country": country,
        }
        for field in M3_FIELDS:
            row[field] = parse_m3_numeric(raw_row[index[field]])
        rows_by_country[country][year] = row
    return rows_by_country


def select_m3_row(m3_country_rows: dict[int, dict], year: int) -> tuple[dict | None, int | None]:
    if not m3_country_rows:
        return None, None
    if year in m3_country_rows:
        return m3_country_rows[year], year
    available_years = sorted(m3_country_rows)
    if year > available_years[-1]:
        source_year = available_years[-1]
        return m3_country_rows[source_year], source_year
    return None, None


def parse_event_countries(country_value) -> list[str]:
    if not country_value or not isinstance(country_value, str):
        return []
    country_value = country_value.strip()
    if country_value in LATAM_COUNTRIES:
        return [country_value]
    parts = []
    for chunk in country_value.replace(" / ", ",").replace(";", ",").split(","):
        part = chunk.strip()
        if part in LATAM_COUNTRIES:
            parts.append(part)
    unique = []
    seen = set()
    for country in parts:
        if country in seen:
            continue
        seen.add(country)
        unique.append(country)
    return unique


def build_annual_event_rollups() -> dict[tuple[str, int], dict[str, int]]:
    if not EVENTS_IN.exists():
        return {}
    payload = json.loads(EVENTS_IN.read_text(encoding="utf-8"))
    events = payload.get("events", [])
    rollups: dict[tuple[str, int], dict[str, int]] = defaultdict(lambda: {
        "sentinel_event_count_y": 0,
        "sentinel_high_salience_event_count_y": 0,
        "sentinel_coup_family_count_y": 0,
        "sentinel_purge_family_count_y": 0,
        "sentinel_domestic_military_role_count_y": 0,
        "sentinel_military_policing_role_count_y": 0,
        "sentinel_exception_rule_militarization_count_y": 0,
    })
    for event in events:
        event_date = str(event.get("event_date") or "")
        if len(event_date) < 4:
            continue
        try:
            year = int(event_date[:4])
        except ValueError:
            continue
        countries = parse_event_countries(event.get("country"))
        if not countries:
            continue
        family = event.get("event_category_family")
        subcategory = event.get("event_subcategory")
        salience = event.get("salience")
        for country in countries:
            row = rollups[(country, year)]
            row["sentinel_event_count_y"] += 1
            if salience == "high":
                row["sentinel_high_salience_event_count_y"] += 1
            if family == "coup":
                row["sentinel_coup_family_count_y"] += 1
            if family == "purge":
                row["sentinel_purge_family_count_y"] += 1
            if subcategory in DOMESTIC_MILITARY_ROLE_SUBCATEGORIES:
                row["sentinel_domestic_military_role_count_y"] += 1
            if subcategory in MILITARY_POLICING_ROLE_SUBCATEGORIES:
                row["sentinel_military_policing_role_count_y"] += 1
            if subcategory in EXCEPTION_RULE_SUBCATEGORIES:
                row["sentinel_exception_rule_militarization_count_y"] += 1
    return rollups


def clamp(value: float, low: float = 0.0, high: float = 100.0) -> float:
    return max(low, min(high, value))


def normalize_unit(value):
    if value is None:
        return None
    try:
        numeric = float(value)
    except (TypeError, ValueError):
        return None
    return clamp(numeric * 100.0)


def normalize_governance_scale(value):
    if value is None:
        return None
    try:
        numeric = float(value)
    except (TypeError, ValueError):
        return None
    return clamp(((numeric + 2.5) / 5.0) * 100.0)


def build_state_capacity(row: dict) -> tuple[float | None, int]:
    components = [
        normalize_unit(row.get("state_authority")),
        normalize_unit(row.get("rule_of_law_vdem")),
        normalize_unit(row.get("judicial_constraints")),
        normalize_unit(row.get("legislative_constraints")),
        normalize_unit(row.get("public_sector_corruption")),
        normalize_unit(row.get("executive_corruption")),
        normalize_unit(row.get("civil_society_participation")),
        normalize_unit(row.get("party_institutionalization")),
    ]
    components[0] = normalize_governance_scale(row.get("state_authority"))
    for key in ("wgi_govt_effectiveness", "wgi_rule_of_law", "wgi_control_corruption"):
        value = row.get(key)
        if value is None:
            continue
        components.append(normalize_governance_scale(value))
    present = [value for value in components if value is not None]
    if not present:
        return None, 0
    return round(sum(present) / len(present), 2), len(present)


def build_year_range(vdem_countries: dict, wb_countries: dict) -> list[int]:
    years: set[int] = set()
    for country in vdem_countries.values():
        for series in country.get("series", {}).values():
            for point in series:
                year = point.get("year")
                if year is not None:
                    years.add(int(year))
    for country in wb_countries.values():
        for key, values in country.items():
            if not key.endswith("_series") or not isinstance(values, list):
                continue
            for point in values:
                year = point.get("year")
                if year is not None:
                    years.add(int(year))
    return sorted(year for year in years if year >= 1960)


def numeric(value):
    if value is None:
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def rounded_delta(current, previous) -> float | None:
    current_num = numeric(current)
    previous_num = numeric(previous)
    if current_num is None or previous_num is None:
        return None
    return round(current_num - previous_num, 3)


def build_country_derived_features(country_rows: list[dict]) -> None:
    last_coup_year: int | None = None
    last_coup_attempt_year: int | None = None
    for idx, row in enumerate(country_rows):
        year = int(row["year"])
        prior_rows = country_rows[:idx]

        row["time_since_last_coup"] = None if last_coup_year is None else year - last_coup_year
        row["time_since_last_coup_attempt"] = (
            None if last_coup_attempt_year is None else year - last_coup_attempt_year
        )

        row["coup_count_5y"] = sum(int(numeric(prev.get("coup_event")) or 0) for prev in prior_rows if year - int(prev["year"]) <= 5)
        row["coup_attempt_count_5y"] = sum(int(numeric(prev.get("coup_attempts")) or 0) for prev in prior_rows if year - int(prev["year"]) <= 5)
        row["coup_count_10y"] = sum(int(numeric(prev.get("coup_event")) or 0) for prev in prior_rows if year - int(prev["year"]) <= 10)
        row["coup_attempt_count_10y"] = sum(int(numeric(prev.get("coup_attempts")) or 0) for prev in prior_rows if year - int(prev["year"]) <= 10)

        if int(numeric(row.get("coup_event")) or 0) > 0:
            last_coup_year = year
        if int(numeric(row.get("coup_attempts")) or 0) > 0:
            last_coup_attempt_year = year

        prev_1 = prior_rows[-1] if len(prior_rows) >= 1 else None
        prev_3 = prior_rows[-3] if len(prior_rows) >= 3 else None

        row["polyarchy_delta_1y"] = rounded_delta(row.get("polyarchy"), prev_1.get("polyarchy") if prev_1 else None)
        row["polyarchy_delta_3y"] = rounded_delta(row.get("polyarchy"), prev_3.get("polyarchy") if prev_3 else None)
        row["mil_exec_delta_1y"] = rounded_delta(row.get("mil_exec"), prev_1.get("mil_exec") if prev_1 else None)
        row["cs_repress_delta_1y"] = rounded_delta(row.get("cs_repress"), prev_1.get("cs_repress") if prev_1 else None)
        row["state_capacity_delta_3y"] = rounded_delta(
            row.get("state_capacity_composite"),
            prev_3.get("state_capacity_composite") if prev_3 else None,
        )
        row["inflation_delta_1y"] = rounded_delta(
            row.get("inflation_consumer_prices_pct"),
            prev_1.get("inflation_consumer_prices_pct") if prev_1 else None,
        )
        row["trade_openness_delta_3y"] = rounded_delta(
            row.get("trade_openness_pct_gdp"),
            prev_3.get("trade_openness_pct_gdp") if prev_3 else None,
        )
        row["oda_received_delta_3y"] = rounded_delta(
            row.get("oda_received_pct_gni"),
            prev_3.get("oda_received_pct_gni") if prev_3 else None,
        )
        row["voter_turnout_delta_1y"] = rounded_delta(
            row.get("voter_turnout"),
            prev_1.get("voter_turnout") if prev_1 else None,
        )

        polyarchy_delta = numeric(row.get("polyarchy_delta_1y"))
        repression_delta = numeric(row.get("cs_repress_delta_1y"))
        inflation_delta = numeric(row.get("inflation_delta_1y"))
        row["regime_shift_flag"] = int(
            int(numeric(row.get("democracy_breakdown")) or 0) > 0
            or (polyarchy_delta is not None and polyarchy_delta <= -0.08)
        )
        row["repression_shift_flag"] = int(repression_delta is not None and repression_delta >= 0.08)
        row["macro_stress_shift_flag"] = int(
            (inflation_delta is not None and abs(inflation_delta) >= 10.0)
            or (numeric(row.get("trade_openness_delta_3y")) is not None and abs(numeric(row.get("trade_openness_delta_3y"))) >= 10.0)
        )


def main(out_dir: Path = OUT_DIR) -> None:
    print(f"Loading V-Dem from {VDEM_IN}")
    vdem_raw = json.loads(VDEM_IN.read_text(encoding="utf-8"))
    vdem_countries = {c["country"]: c for c in vdem_raw["countries"]}

    print(f"Loading World Bank from {WB_IN}")
    wb_raw = json.loads(WB_IN.read_text(encoding="utf-8"))
    wb_countries = {c["country"]: c for c in wb_raw["countries"]}

    print(f"Loading M3 from {M3_IN}")
    m3_countries = load_m3_rows()

    print(f"Loading annual event rollups from {EVENTS_IN}")
    event_rollups = build_annual_event_rollups()

    years = build_year_range(vdem_countries, wb_countries)

    rows: list[dict] = []

    all_countries = sorted(set(list(vdem_countries.keys()) + list(wb_countries.keys())))

    for country in all_countries:
        iso3 = ISO3.get(country, "")
        vdem = vdem_countries.get(country)
        wb   = wb_countries.get(country)
        m3_country_rows = m3_countries.get(country, {})

        # Build V-Dem year → value lookup
        vdem_series: dict[str, dict[int, float | None]] = {}
        if vdem and "series" in vdem:
            for field in VDEM_SERIES:
                if field in vdem["series"]:
                    vdem_series[field] = {pt["year"]: pt["value"] for pt in vdem["series"][field]}

        # Build WB year → value lookup
        wb_series: dict[str, dict[int, float | None]] = {}
        if wb:
            for wb_field, col_name in WB_SERIES.items():
                if wb_field in wb:
                    wb_series[col_name] = {int(pt["year"]): pt["value"] for pt in wb[wb_field]}

        for year in years:
            m3_row, m3_source_year = select_m3_row(m3_country_rows, year)
            annual_event_rollup = event_rollups.get((country, year), {})
            row: dict = {
                "country":    country,
                "iso3":       iso3,
                "year":       year,
                # V-Dem indicators
                **{f: vdem_series.get(f, {}).get(year) for f in VDEM_SERIES},
                # WB indicators
                **{col: wb_series.get(col, {}).get(year) for col in WB_SERIES.values()},
                # M3 structural (annual where available, forward-filled after latest observation)
                "m3_source_year":              m3_source_year,
                "m3_observed_year":            int(m3_source_year == year) if m3_source_year is not None else 0,
                "m3_conscription":             m3_row.get("com_mil_serv") if m3_row else None,
                "m3_conscription_gender":      m3_row.get("com_mil_serv_gen") if m3_row else None,
                "m3_conscription_dur_max":     m3_row.get("com_mil_serv_dur_max") if m3_row else None,
                "m3_conscription_dur_min":     m3_row.get("com_mil_serv_dur_min") if m3_row else None,
                "m3_alt_civil_service":        m3_row.get("alt_civ_serv") if m3_row else None,
                "m3_mil_origin":               m3_row.get("mil_origin") if m3_row else None,
                "m3_mil_leader":               m3_row.get("mil_leader") if m3_row else None,
                "m3_mil_mod":                  m3_row.get("mil_mod") if m3_row else None,
                "m3_mil_veto":                 m3_row.get("mil_veto") if m3_row else None,
                "m3_mil_repress":              m3_row.get("mil_repress") if m3_row else None,
                "m3_mil_repress_count":        m3_row.get("mil_repress_count") if m3_row else None,
                "m3_mil_impunity":             m3_row.get("mil_impun") if m3_row else None,
                "m3_mil_crime_police":         m3_row.get("milpol_crime") if m3_row else None,
                "m3_mil_law_enforcement":      m3_row.get("milpol_law") if m3_row else None,
                "m3_mil_peace_order":          m3_row.get("milpol_peace") if m3_row else None,
                "m3_mil_police_overlap":       m3_row.get("mil_police") if m3_row else None,
                "m3_mil_eco":                  m3_row.get("mil_eco_dummy") if m3_row else None,
                "m3_mil_eco_own":              m3_row.get("mil_eco_own") if m3_row else None,
                "m3_mil_eco_share":            m3_row.get("mil_eco_share") if m3_row else None,
                "m3_mil_eco_dom":              m3_row.get("mil_eco_dom") if m3_row else None,
                "m3_milex_gdp":                m3_row.get("milex_gdp") if m3_row else None,
                "m3_milex_healthexp":          m3_row.get("milex_healthexp") if m3_row else None,
                "m3_pers_to_pop":              m3_row.get("pers_to_pop") if m3_row else None,
                "m3_pers_to_phy":              m3_row.get("pers_to_phy") if m3_row else None,
                "m3_reserve_pop":              m3_row.get("reserve_pop") if m3_row else None,
                "m3_hwi":                      m3_row.get("hwi") if m3_row else None,
                # Annualized SENTINEL event rollups
                "sentinel_event_count_y": annual_event_rollup.get("sentinel_event_count_y", 0),
                "sentinel_high_salience_event_count_y": annual_event_rollup.get("sentinel_high_salience_event_count_y", 0),
                "sentinel_coup_family_count_y": annual_event_rollup.get("sentinel_coup_family_count_y", 0),
                "sentinel_purge_family_count_y": annual_event_rollup.get("sentinel_purge_family_count_y", 0),
                "sentinel_domestic_military_role_count_y": annual_event_rollup.get("sentinel_domestic_military_role_count_y", 0),
                "sentinel_military_policing_role_count_y": annual_event_rollup.get("sentinel_military_policing_role_count_y", 0),
                "sentinel_exception_rule_militarization_count_y": annual_event_rollup.get("sentinel_exception_rule_militarization_count_y", 0),
            }
            state_capacity, state_capacity_coverage = build_state_capacity(row)
            row["state_capacity_composite"] = state_capacity
            row["state_capacity_coverage"] = state_capacity_coverage
            rows.append(row)

    rows.sort(key=lambda r: (r["country"], r["year"]))
    grouped: dict[str, list[dict]] = {}
    for row in rows:
        grouped.setdefault(row["country"], []).append(row)
    for country_rows in grouped.values():
        build_country_derived_features(country_rows)
    print(f"  {len(rows)} country-year rows built ({len(all_countries)} countries × {len(years)} years)")

    out_dir.mkdir(parents=True, exist_ok=True)

    # ── Write JSON ─────────────────────────────────────────────
    json_path = out_dir / "country_year.json"
    json_path.write_text(
        json.dumps({
            "generated": datetime.now(UTC).isoformat(),
            "schema_version": "1.0",
            "count": len(rows),
            "columns": list(rows[0].keys()) if rows else [],
            "rows": rows,
        }, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    print(f"JSON written: {json_path}")

    # ── Write CSV ──────────────────────────────────────────────
    csv_path = out_dir / "country_year.csv"
    if rows:
        with csv_path.open("w", newline="", encoding="utf-8") as f:
            writer = csv.DictWriter(f, fieldnames=list(rows[0].keys()))
            writer.writeheader()
            writer.writerows(rows)
        print(f"CSV written: {csv_path}")

    # ── Coverage summary ───────────────────────────────────────
    from collections import Counter
    non_null = Counter()
    for r in rows:
        for k, v in r.items():
            if v is not None and k not in ("country", "iso3", "year"):
                non_null[k] += 1
    print("\nField coverage (non-null rows):")
    for field, n in sorted(non_null.items(), key=lambda x: -x[1]):
        pct = n / len(rows) * 100
        print(f"  {field:40s} {n:4d} / {len(rows)} ({pct:.0f}%)")


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Build SENTINEL country-year join table")
    parser.add_argument("--output-dir", type=Path, default=OUT_DIR)
    args = parser.parse_args()
    main(args.output_dir)
