"""
clean_acled_index.py
Reads the ACLED Conflict Index Excel workbook and writes a SENTINEL-ready
JSON to data/cleaned/acled_index.json.

Usage:
    python scripts/clean_acled_index.py
    # Reads data/raw/ACLED_Conflict_Index_2025.xlsx relative to repo root.

Requires openpyxl:
    pip install openpyxl
"""

import json
import os
import sys
from datetime import datetime, timezone

# ---------------------------------------------------------------------------
# Ensure openpyxl is available
# ---------------------------------------------------------------------------
try:
    import openpyxl
except ImportError:
    import subprocess
    print("openpyxl not found — installing …")
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl", "-q"])
    import openpyxl  # noqa: F811

# ---------------------------------------------------------------------------
# Paths
# ---------------------------------------------------------------------------
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
REPO_ROOT   = os.path.join(SCRIPT_DIR, "..")
RAW_XLSX    = os.path.join(REPO_ROOT, "data", "raw", "ACLED_Conflict_Index_2025.xlsx")
OUT_DIR     = os.path.join(REPO_ROOT, "data", "cleaned")
OUT_JSON    = os.path.join(OUT_DIR, "acled_index.json")

# ---------------------------------------------------------------------------
# ACLED country name  →  SENTINEL canonical name
# ---------------------------------------------------------------------------
NAME_MAP: dict[str, str] = {
    "Brazil":              "Brazil",
    "Colombia":            "Colombia",
    "Mexico":              "Mexico",
    "Venezuela":           "Venezuela",
    "Argentina":           "Argentina",
    "Peru":                "Peru",
    "Chile":               "Chile",
    "Ecuador":             "Ecuador",
    "Bolivia":             "Bolivia",
    "Honduras":            "Honduras",
    "Nicaragua":           "Nicaragua",
    "Guatemala":           "Guatemala",
    "El Salvador":         "El Salvador",
    "Paraguay":            "Paraguay",
    "Uruguay":             "Uruguay",
    "Cuba":                "Cuba",
    "Haiti":               "Haiti",
    "Dominican Republic":  "Dominican Republic",
    "Panama":              "Panama",
    "Costa Rica":          "Costa Rica",
    "Jamaica":             "Jamaica",
    "Trinidad and Tobago": "Trinidad and Tobago",
    "Guyana":              "Guyana",
    "Suriname":            "Suriname",
    "Belize":              "Belize",
}

# ---------------------------------------------------------------------------
# Expected column headers in the Results sheet (exact strings)
# ---------------------------------------------------------------------------
EXPECTED_HEADERS = [
    "Country",
    "Index Level",
    "Index Ranking",
    "Deadliness Ranking",
    "Diffusion Ranking",
    "Danger Ranking",
    "Fragmentation Ranking",
    "Deadliness Value",
    "Diffusion Value",
    "Danger Value",
    "Fragmentation Value",
]

# Header  →  output key
HEADER_KEY_MAP = {
    "Country":               "country",
    "Index Level":           "index_level",
    "Index Ranking":         "index_ranking",
    "Deadliness Ranking":    "deadliness_ranking",
    "Diffusion Ranking":     "diffusion_ranking",
    "Danger Ranking":        "danger_ranking",
    "Fragmentation Ranking": "fragmentation_ranking",
    "Deadliness Value":      "deadliness_value",
    "Diffusion Value":       "diffusion_value",
    "Danger Value":          "danger_value",
    "Fragmentation Value":   "fragmentation_value",
}

# Keys that should be stored as int
INT_KEYS = {
    "index_ranking", "deadliness_ranking", "diffusion_ranking",
    "danger_ranking", "fragmentation_ranking",
    "deadliness_value", "danger_value", "fragmentation_value",
}

# Keys that should be stored as float
FLOAT_KEYS = {"diffusion_value"}


def coerce(key: str, val):
    """Type-coerce a cell value for the output key."""
    if val is None or str(val).strip() == "":
        return None
    if key in INT_KEYS:
        try:
            return int(val)
        except (ValueError, TypeError):
            return None
    if key in FLOAT_KEYS:
        try:
            return float(val)
        except (ValueError, TypeError):
            return None
    # Default: return as-is (strings like index_level, country)
    return val


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    if not os.path.exists(RAW_XLSX):
        print(f"ERROR: raw file not found: {RAW_XLSX}")
        sys.exit(1)

    os.makedirs(OUT_DIR, exist_ok=True)
    print(f"Reading {RAW_XLSX} …")

    wb = openpyxl.load_workbook(RAW_XLSX, read_only=True, data_only=True)

    if "Results" not in wb.sheetnames:
        print(f"ERROR: 'Results' sheet not found. Available sheets: {wb.sheetnames}")
        sys.exit(1)

    ws = wb["Results"]
    rows = list(ws.iter_rows(values_only=True))

    if not rows:
        print("ERROR: Results sheet is empty.")
        sys.exit(1)

    # Parse header row
    raw_headers = [str(h).strip() if h is not None else "" for h in rows[0]]
    missing = [h for h in EXPECTED_HEADERS if h not in raw_headers]
    if missing:
        print(f"  WARNING: expected columns not found in sheet: {missing}")
        print(f"  Actual headers: {raw_headers}")

    # Build column index map: output_key → column index
    col_idx: dict[str, int] = {}
    for h, key in HEADER_KEY_MAP.items():
        try:
            col_idx[key] = raw_headers.index(h)
        except ValueError:
            pass  # already warned above

    print(f"  Total rows in Results sheet (excl. header): {len(rows) - 1}")

    countries_out = []
    found_names = []
    skipped = 0

    for row in rows[1:]:
        if not any(row):
            continue  # skip blank rows

        acled_name_raw = row[col_idx["country"]] if "country" in col_idx else None
        if acled_name_raw is None:
            skipped += 1
            continue
        acled_name = str(acled_name_raw).strip()

        if acled_name not in NAME_MAP:
            skipped += 1
            continue

        sentinel_name = NAME_MAP[acled_name]
        found_names.append(sentinel_name)

        record: dict = {"country": sentinel_name}
        for key, idx in col_idx.items():
            if key == "country":
                continue
            raw_val = row[idx] if idx < len(row) else None
            record[key] = coerce(key, raw_val)

        countries_out.append(record)

    # Sort by index_ranking ascending (None last)
    countries_out.sort(key=lambda r: (r.get("index_ranking") is None, r.get("index_ranking") or 9999))

    print(f"\n  LatAm countries found ({len(found_names)}):")
    for name in sorted(found_names):
        idx_rank = next((r["index_ranking"] for r in countries_out if r["country"] == name), "?")
        idx_level = next((r["index_level"] for r in countries_out if r["country"] == name), "?")
        print(f"    {name:<25}  ranking={idx_rank}  level={idx_level}")

    not_found = sorted(set(NAME_MAP.values()) - set(found_names))
    if not_found:
        print(f"\n  LatAm countries NOT in ACLED Index ({len(not_found)}): {not_found}")
        print("  (These are likely absent from the index because conflict intensity is very low.)")

    # ---------------------------------------------------------------------------
    # Write JSON
    # ---------------------------------------------------------------------------
    payload = {
        "updated": datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ"),
        "source":  "ACLED Conflict Index December 2025",
        "index_levels": ["Low", "Medium", "High", "Very High", "Extreme"],
        "columns": {
            "index_level":           "Overall conflict severity: Low/Medium/High/Very High/Extreme",
            "index_ranking":         "Global ranking (1=most conflict)",
            "deadliness_value":      "Total fatalities in past 12 months",
            "diffusion_value":       "Share of admin units with political violence events",
            "danger_value":          "Fatalities from targeted civilian violence",
            "fragmentation_value":   "Number of distinct armed groups active",
        },
        "countries": countries_out,
    }

    with open(OUT_JSON, "w", encoding="utf-8") as f:
        json.dump(payload, f, ensure_ascii=False, indent=2)

    print(f"\n  → {OUT_JSON}  ({len(countries_out)} countries)")
    print("\nDone.")


if __name__ == "__main__":
    main()
