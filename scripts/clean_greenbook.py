"""
clean_greenbook.py
Cleans the USAID US Foreign Aid Greenbook XLSX and writes SENTINEL-ready outputs
to data/cleaned/greenbook.json and data/cleaned/greenbook.csv.

Usage:
    python scripts/clean_greenbook.py
    # Reads data/raw/us_foreignaid_greenbook.xlsx relative to the repo root.
"""

import csv
import json
import os
import sys
from datetime import datetime, timezone
from collections import defaultdict

# ---------------------------------------------------------------------------
# Paths (resolved relative to this script file so it works from any cwd)
# ---------------------------------------------------------------------------
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
REPO_ROOT  = os.path.join(SCRIPT_DIR, "..")
RAW_XLSX   = os.path.join(REPO_ROOT, "data", "raw", "us_foreignaid_greenbook.xlsx")
OUT_DIR    = os.path.join(REPO_ROOT, "data", "cleaned")
OUT_JSON   = os.path.join(OUT_DIR, "greenbook.json")
OUT_CSV    = os.path.join(OUT_DIR, "greenbook.csv")

# ---------------------------------------------------------------------------
# Greenbook country name  →  SENTINEL canonical name
# ---------------------------------------------------------------------------
NAME_MAP: dict[str, str] = {
    "Argentina":                        "Argentina",
    "Belize":                           "Belize",
    "Bolivia":                          "Bolivia",
    "Brazil":                           "Brazil",
    "Chile":                            "Chile",
    "Colombia":                         "Colombia",
    "Costa Rica":                       "Costa Rica",
    "Cuba":                             "Cuba",
    "Dominican Republic":               "Dominican Republic",
    "Ecuador":                          "Ecuador",
    "El Salvador":                      "El Salvador",
    "Guatemala":                        "Guatemala",
    "Guyana":                           "Guyana",
    "Haiti":                            "Haiti",
    "Honduras":                         "Honduras",
    "Jamaica":                          "Jamaica",
    "Mexico":                           "Mexico",
    "Nicaragua":                        "Nicaragua",
    "Panama":                           "Panama",
    "Paraguay":                         "Paraguay",
    "Peru":                             "Peru",
    "Suriname":                         "Suriname",
    "Trinidad and Tobago":              "Trinidad & Tobago",
    "Trinidad & Tobago":                "Trinidad & Tobago",
    "Uruguay":                          "Uruguay",
    "Venezuela":                        "Venezuela",
    # Regional aggregates
    "Latin America and the Caribbean":  "Regional - LAC",
    "Central America":                  "Regional - Central America",
    "Caribbean":                        "Regional - Caribbean",
    "South America":                    "Regional - South America",
}

# Normalised set of Greenbook names to match against
GREENBOOK_NAMES: set[str] = set(NAME_MAP.keys())


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def safe_float(val) -> float:
    """Return float, or 0.0 for blanks / non-numeric values."""
    if val is None:
        return 0.0
    s = str(val).strip()
    if s == "" or s.lower() in ("na", "nan", "none", "."):
        return 0.0
    try:
        return float(s)
    except ValueError:
        return 0.0


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    try:
        import openpyxl
    except ImportError:
        print("ERROR: openpyxl is required.  Run: pip install openpyxl")
        sys.exit(1)

    if not os.path.exists(RAW_XLSX):
        print(f"ERROR: raw XLSX not found: {RAW_XLSX}")
        sys.exit(1)

    os.makedirs(OUT_DIR, exist_ok=True)
    print(f"Reading {RAW_XLSX} …")

    wb = openpyxl.load_workbook(RAW_XLSX, read_only=True, data_only=True)
    ws = wb.active

    # Identify header row and column indices
    header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    headers = [str(h).strip() if h is not None else "" for h in header_row]

    def col(name: str) -> int:
        """Return 0-based index of a column by name (case-insensitive prefix match)."""
        name_l = name.lower()
        for i, h in enumerate(headers):
            if h.lower() == name_l:
                return i
        # Fallback: substring match
        for i, h in enumerate(headers):
            if name_l in h.lower():
                return i
        raise KeyError(f"Column not found: {name!r}  (headers: {headers})")

    try:
        i_year     = col("Fiscal Year")
        i_country  = col("Country")
        i_category = col("Assistance Category")
        i_constant = col("Obligations (Constant Dollars)")
    except KeyError as e:
        print(f"ERROR: {e}")
        wb.close()
        sys.exit(1)

    print(f"  Columns: year={i_year}, country={i_country}, "
          f"category={i_category}, constant_usd={i_constant}")

    # country_sentinel  →  year (int)  →  {"economic": float, "military": float}
    data: dict[str, dict[int, dict[str, float]]] = defaultdict(
        lambda: defaultdict(lambda: {"economic": 0.0, "military": 0.0})
    )

    row_count   = 0
    kept_count  = 0
    skipped_cat = 0

    for row in ws.iter_rows(min_row=2, values_only=True):
        row_count += 1

        raw_country  = row[i_country]
        raw_year     = row[i_year]
        raw_category = row[i_category]
        raw_amount   = row[i_constant]

        if raw_country is None:
            continue

        country_str = str(raw_country).strip()
        if country_str not in GREENBOOK_NAMES:
            continue

        try:
            year = int(float(str(raw_year)))
        except (ValueError, TypeError):
            continue

        amount = safe_float(raw_amount)
        category = str(raw_category).strip().lower() if raw_category else ""

        sentinel_name = NAME_MAP[country_str]

        if "economic" in category:
            data[sentinel_name][year]["economic"] += amount
        elif "military" in category:
            data[sentinel_name][year]["military"] += amount
        else:
            skipped_cat += 1
            continue

        kept_count += 1

    wb.close()
    print(f"  Scanned {row_count:,} rows; kept {kept_count:,} LatAm rows "
          f"({skipped_cat} skipped — unknown category)")

    # ---------------------------------------------------------------------------
    # Build per-country objects
    # ---------------------------------------------------------------------------
    countries_out = []
    csv_rows      = []

    for sentinel_name, year_data in data.items():
        series = []
        total_economic = 0.0
        total_military = 0.0
        peak_year      = None
        peak_value     = 0.0

        for year in sorted(year_data.keys()):
            eco  = year_data[year]["economic"]
            mil  = year_data[year]["military"]
            tot  = eco + mil
            if tot <= 0:
                continue

            series.append({
                "year":     year,
                "economic": round(eco, 2),
                "military": round(mil, 2),
                "total":    round(tot, 2),
            })

            total_economic += eco
            total_military += mil

            if tot > peak_value:
                peak_value = tot
                peak_year  = year

        total_all = total_economic + total_military
        if total_all <= 0:
            continue

        countries_out.append({
            "country":        sentinel_name,
            "total_all_years": round(total_all, 2),
            "total_economic":  round(total_economic, 2),
            "total_military":  round(total_military, 2),
            "peak_year":       peak_year,
            "peak_value":      round(peak_value, 2),
            "series":          series,
        })

        for pt in series:
            csv_rows.append({
                "country":  sentinel_name,
                "year":     pt["year"],
                "economic": pt["economic"],
                "military": pt["military"],
                "total":    pt["total"],
            })

    # Sort countries by total_all_years descending
    countries_out.sort(key=lambda x: x["total_all_years"], reverse=True)

    # ---------------------------------------------------------------------------
    # Write JSON
    # ---------------------------------------------------------------------------
    payload = {
        "updated": datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ"),
        "source":  "USAID US Foreign Aid Greenbook (Congressional Budget Justification)",
        "note":    "Constant dollar obligations. Fiscal years 1946-2023.",
        "columns": {
            "year":     "US Fiscal Year",
            "economic": "Economic assistance obligations (constant USD)",
            "military": "Military assistance obligations (constant USD)",
            "total":    "Total obligations (constant USD)",
        },
        "countries": countries_out,
    }

    with open(OUT_JSON, "w", encoding="utf-8") as f:
        json.dump(payload, f, ensure_ascii=False, indent=2)
    print(f"\n  → {OUT_JSON}  ({len(countries_out)} countries)")

    # ---------------------------------------------------------------------------
    # Write CSV
    # ---------------------------------------------------------------------------
    csv_fieldnames = ["country", "year", "economic", "military", "total"]
    with open(OUT_CSV, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=csv_fieldnames, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(csv_rows)
    print(f"  → {OUT_CSV}  ({len(csv_rows)} rows)")

    # ---------------------------------------------------------------------------
    # Print top 10
    # ---------------------------------------------------------------------------
    print("\nTop 10 LatAm recipients (constant USD, all years):")
    print(f"  {'Country':<35}  {'Total':>18}  {'Economic':>18}  {'Military':>18}")
    print(f"  {'-'*35}  {'-'*18}  {'-'*18}  {'-'*18}")
    for c in countries_out[:10]:
        print(
            f"  {c['country']:<35}  "
            f"${c['total_all_years']:>17,.0f}  "
            f"${c['total_economic']:>17,.0f}  "
            f"${c['total_military']:>17,.0f}"
        )

    print("\nDone.")


if __name__ == "__main__":
    main()
